"""
bionetverification
Modified by: Michelle Aluf-Medina
"""

import logging
from openpyxl import load_workbook as loadwb
import sat
import ssp
import ec
import gn
import miscfunctions as misc
import os


def manual_menu():
    """
    MAIN
    -----
    This is the main body of the script. All functions are called from here.

    LOGGING
    --------
    Setup log file and logging format
    """
    log_dir = 'Logs'
    if not os.path.exists(log_dir):
        os.mkdir(log_dir)
    log_path = misc.file_name_cformat(os.path.join(os.getcwd(), log_dir, 'log_{0}.log'))
    fmt = ('%(asctime)s\t%(module)s\t%(funcName)s\t-\t\t%(levelname)s:\t%(message)s')
    logging.basicConfig(filename=log_path, level=logging.DEBUG, format=fmt,
                        datefmt='%Y-%m-%d %H:%M:%S')

    """
    CURRENT WORKING DIRECTORY
    --------------------------
    Make a directory for output files
    Change the working directory to this new directory
    """
    logging.info('Change working directory')
    dir_name = misc.file_name_cformat('bionetverification_out_{0}')
    os.mkdir(dir_name)
    cwd = os.getcwd()
    template_dir = cwd + '/Template/'
    input_dir = cwd + '/Inputs/'
    new_cwd = os.path.join(cwd, dir_name)
    os.chdir(new_cwd)
    cwd = os.getcwd()
    print('Current working directory:\n' + cwd)
    logging.info('New working directory:' + cwd)

    """
    PROBLEM TYPE SELECTION
    -----------------------
    Get the problem type being looked at
        [0] General Circuit
        [1] SSP
        [2] ExCov
        [3] SAT
        [4] Quit
    """
    problem_type = ''
    while problem_type != 4:  # While not quit
        print('Bionetverification\n')
        logging.info('Bionetverification')
        logging.info('Printing problem selection menu')
        misc.print_menu()
        problem_type = misc.int_input()
        logging.info('Selected option: ' + str(problem_type))

        """
        General Network SELECTED
        """
        if problem_type == 0:
            """
            Get and Parse General Network from input file
            --------------------------------------
            """
            # While filename is not in Inputs
            gn_pstr = 'Please enter Network filename: '
            gn_smv_fn = misc.input_exists(input_dir, gn_pstr)
            depth, split_junc, force_down_junc = misc.read_gn(gn_smv_fn)

            # Setup worksheet for data recording
            gn_wb = loadwb(template_dir + 'GN_Template.xlsx')
            gn_wb_num_of_sheets = len(gn_wb.sheetnames)
            gn_xl_fn = misc.file_name_cformat('gn_{0}.xlsx')
            gn_wb.save(gn_xl_fn)

            """
            Run Model Checker
            Run new specs (csum and nsum for whole network)
            ------------------
            """

            str_modc = misc.modcheck_select()
            if str_modc == 'NuSMV':
                # Add another worksheet based on the template
                s_source = gn_wb['SINGLE_Template']
                gn_s_ws = gn_wb.copy_worksheet(s_source)
                gn_s_ws.title = ('GN_GenSpec')
                gn_wb.save(gn_xl_fn)

                # generate smv file
                gn_smv_fn = f'GN_depth_{depth}.smv'
                gn.smv_gen(gn_smv_fn, depth, split_junc, force_down_junc)

                # run smv file
                gn.run_nusmv_gn([gn_smv_fn], gn_wb, gn_s_ws, gn_xl_fn, str_modchecker=str_modc, depth=[depth])

            elif str_modc == 'prism':
                logging.info('Printing GC menu')
                ssp.print_ssp_menu(str_modc)
                logging.info('Printed GC menu.')
                # Get user input for menu selection
                valid_opt = -1
                ssp_opt = None
                while valid_opt == -1:
                    ssp_opt = misc.int_input()
                    if ssp_opt in range(1, 4):
                        valid_opt = 1
                        print('Selected option ', str(ssp_opt))
                        logging.info('Selected option: ' + str(ssp_opt))
                    else:
                        print('Invalid option selected. '
                              'Please select a number 1-4.')
                        logging.info('Invalid option selected.')

                # choose mu
                mu_user_input = misc.prism_set_mu()

                # Add another worksheet based on the template
                s_source = gn_wb['Prism_Template']
                gn_s_ws = gn_wb.copy_worksheet(s_source)
                gn_s_ws.title = ('GN_GenSpec')
                gn_wb.save(gn_xl_fn)
                gn_smv_fn = f'GN_depth_{depth}_mu_{mu_user_input}.pm'

                # generate prism file
                gn.prism_gen(gn_smv_fn, depth, split_junc, force_down_junc, mu=mu_user_input)
                if mu_user_input > .0:
                    gn_smv_fn_no_error = f'GN_depth_{depth}_mu_0.pm'
                    gn.prism_gen(gn_smv_fn_no_error, depth, split_junc, force_down_junc, mu=mu_user_input)
                    gn_smv_fn_arr = [gn_smv_fn_no_error, gn_smv_fn]
                else:
                    gn_smv_fn_arr = [gn_smv_fn]
                gn.gen_prism_spec('spec_gn.pctl')

                # run prism file
                gn.run_prism_gn(gn_smv_fn_arr, gn_wb, gn_s_ws, gn_xl_fn, str_modchecker=str_modc, depth=[depth], spec_number=ssp_opt, mu=mu_user_input)

            if len(gn_wb.sheetnames) > gn_wb_num_of_sheets:
                gn_wb.remove(gn_wb['SINGLE_Template'])
                gn_wb.remove(gn_wb['Prism_Template'])
                gn_wb.save(gn_xl_fn)
                logging.info('Output Excel file is: ' + gn_xl_fn)
            else:
                logging.info('There was no action in file: ' + gn_xl_fn)
                logging.info('delete file: ' + gn_xl_fn)
                os.remove(gn_xl_fn)
            logging.info('Closing workbook')
            gn_wb.close()
        """
        SSP SELECTED
        """
        if problem_type == 1:
            """
            Get and Parse SSP sets from input file
            --------------------------------------
            """
            # While filename is not in Inputs
            ssp_pstr = 'Please enter SSP problems filename: '
            ssp_fn = misc.input_exists(input_dir, ssp_pstr)
            ssp_arr, num_sets = ssp.read_ssp(ssp_fn)

            # Setup worksheet for data recording
            ssp_wb = loadwb(template_dir + 'SSP_Template.xlsx')
            ssp_wb_num_of_sheets = len(ssp_wb.sheetnames)
            ssp_xl_fn = misc.file_name_cformat('SSP_{0}.xlsx')
            ssp_wb.save(ssp_xl_fn)

            # Pick a model checker (NuSMV, nuXmv or PRISM)
            str_modc = misc.modcheck_select()
            ssp_opt = ''

            if str_modc == "NuSMV" or str_modc == "nuXmv":
                """
                Generate smv files
                """
                # Use specification per output
                ssp_smv, ssp_smv_nt = ssp.smv_gen(ssp_arr, str_modc)

                # Use new specifications (csum and nsum for whole network)
                ssp_smv_new, ssp_smv_nt_new = ssp.smv_gen_newspec(ssp_arr, str_modc)

                while ssp_opt != 4:
                    logging.info('Printing SSP menu')
                    ssp.print_ssp_menu(str_modc)
                    logging.info('Printed SSP menu.')
                    # Get user input for menu selection
                    valid_opt = -1
                    while valid_opt == -1:
                        ssp_opt = misc.int_input()
                        if ssp_opt in range(1, 5):
                            valid_opt = 1
                            print('Selected option ', str(ssp_opt))
                            logging.info('Selected option: ' + str(ssp_opt))
                        else:
                            print('Invalid option selected. '
                                  'Please select a number 1-4.')
                            logging.info('Invalid option selected.')

                    # If selected bulk run
                    if ssp_opt == 1:
                        """
                        Run NuSMV
                        Bulk run specs for per output specs
                        ------------------
                        """
                        # Add another worksheet based on the template
                        a_source = ssp_wb['ALL_Template']
                        ssp_a_ws = ssp_wb.copy_worksheet(a_source)
                        ssp_a_ws.title = ('Bulk_OutSpec')
                        ssp_wb.save(ssp_xl_fn)

                        # Run NuSMV and get output filename for specification
                        ssp.run_nusmv_all(ssp_arr, ssp_smv, ssp_smv_nt, ssp_wb, ssp_a_ws, ssp_xl_fn, str_modc)

                    # If selected individual out run
                    elif ssp_opt == 2:
                        """
                        Run NuSMV
                        Run each per output spec individually
                        ------------------
                        """
                        # Add another worksheet based on the template
                        s_source = ssp_wb['SINGLE_Template']
                        ssp_s_ws = ssp_wb.copy_worksheet(s_source)
                        ssp_s_ws.title = ('Single_OutSpec')
                        ssp_wb.save(ssp_xl_fn)

                        # Run NuSMV and get outputs for each individual specification
                        ssp.run_nusmv_single(ssp_arr, ssp_smv, ssp_smv_nt, ssp_wb, ssp_s_ws, ssp_xl_fn, str_modc)

                    # If selected general specifications
                    elif ssp_opt == 3:
                        """
                        Run NuSMV
                        Run new specs (csum and nsum for whole network)
                        ------------------
                        """
                        # Add another worksheet based on the template
                        s_source = ssp_wb['NewSpec_Template']
                        ssp_s_ws = ssp_wb.copy_worksheet(s_source)
                        ssp_s_ws.title = ('SSP_GenSpec')
                        ssp_wb.save(ssp_xl_fn)

                        # Run NuSMV and get outputs for each individual specification
                        ssp.run_nusmv_newspec(ssp_arr, ssp_smv_new, ssp_smv_nt_new, ssp_wb, ssp_s_ws, ssp_xl_fn,
                                              str_modc)

            elif str_modc == "prism":
                while ssp_opt != 3:
                    logging.info('Printing SSP menu')
                    ssp.print_ssp_menu(str_modc)
                    logging.info('Printed SSP menu.')
                    # Get user input for menu selection
                    valid_opt = -1
                    while valid_opt == -1:
                        ssp_opt = misc.int_input()
                        if ssp_opt in range(1, 4):
                            valid_opt = 1
                            print('Selected option ', str(ssp_opt))
                            logging.info('Selected option: ' + str(ssp_opt))
                        else:
                            print('Invalid option selected. '
                                  'Please select a number 1-3.')
                            logging.info('Invalid option selected.')

                    if ssp_opt == 1 or ssp_opt == 2:
                        """
                        Generate prism files
                        """
                        # Use specification per output
                        mu_user_input = misc.prism_set_mu()
                        ssp_prism_nt = ssp.prism_gen(ssp_arr, mu_user_input)

                        """
                        Run Prism
                        ------------------
                        """
                        # Add another worksheet based on the template
                        s_source = ssp_wb['Prism_Template']
                        ssp_s_ws = ssp_wb.copy_worksheet(s_source)
                        ssp_s_ws.title = 'SSP_Prism_Results'
                        ssp_wb.save(ssp_xl_fn)

                        # Run Prism and get outputs for each individual specification
                        ssp.run_prism(ssp_arr, ssp_prism_nt, ssp_wb, ssp_s_ws, ssp_xl_fn, str_modc, ssp_opt)

            # If selected return to main
            """
            Finished running SSP problems
            Remove template sheets and close file
            In case there was no action - delete the file
            """
            if len(ssp_wb.sheetnames) > ssp_wb_num_of_sheets:
                ssp_wb.remove(ssp_wb['ALL_Template'])
                ssp_wb.remove(ssp_wb['SINGLE_Template'])
                ssp_wb.remove(ssp_wb['NewSpec_Template'])
                ssp_wb.remove(ssp_wb['Prism_Template'])
                ssp_wb.save(ssp_xl_fn)
                logging.info('Output Excel file is: ' + ssp_xl_fn)
            else:
                logging.info('There was no action in file: ' + ssp_xl_fn)
                logging.info('delete file: ' + ssp_xl_fn)
                os.remove(ssp_xl_fn)
            logging.info('Closing workbook')
            ssp_wb.close()

            # create MATLAB file
            logging.info('Create MATLAB file')
            misc.ssp_create_m_file(Su=ssp_arr)

        """
        ExCov SELECTED
        """
        if problem_type == 2:
            """
            Get and Parse ExCov universes and sets from input file
            --------------------------------------
            """
            # While filename is not in Inputs
            excov_pstr = 'Please enter ExCov problems filename: '
            excov_fn = misc.input_exists(input_dir, excov_pstr)
            universes, subsets_arrays, num_probs = ec.read_ec(excov_fn)

            # Pick a model checker (NuSMV, nuXmv or PRISM)
            str_modc = misc.modcheck_select()

            # Setup worksheet for data recording
            ec_wb = loadwb(template_dir + 'EC_Template.xlsx')
            ec_xl_fn = misc.file_name_cformat('ExCov_{0}.xlsx')
            ec_wb.save(ec_xl_fn)

            if str_modc == "NuSMV" or str_modc == "nuXmv":
                """
                Generate smv files
                """
                ec_smv, ec_smv_nt, ec_outputs, max_sums = ec.smv_gen(universes, subsets_arrays, bit_mapping=True)

                """
                Run NuSMV
                Single spec for exact cover output
                ------------------
                """
                # Add another worksheet based on the template
                source = ec_wb['EC_Template']
                ec_ws = ec_wb.copy_worksheet(source)
                ec_ws.title = 'ExCov'
                ec_wb.save(ec_xl_fn)

                # Run NuSMV and get outputs for each individual specification
                ec.run_nusmv(universes, subsets_arrays, ec_outputs, ec_smv, ec_smv_nt, ec_wb, ec_ws, ec_xl_fn, str_modc, max_sums)

            elif str_modc == "prism":

                """
                Generate prism files
                """

                # choose type of check
                ec_opt = ec.ec_prism_menu()

                while ec_opt != 3:
                    mu_user_input = misc.prism_set_mu()
                    ec_prism_nt, ec_outputs, max_sums = ec.prism_gen(universes, subsets_arrays, mu_user_input,
                                                                     bit_mapping=True)

                    """
                    Run prism
                    Single spec for exact cover output
                    ------------------
                    """
                    # Add another worksheet based on the template
                    source = ec_wb['EC_Prism_Template']
                    ec_ws = ec_wb.copy_worksheet(source)
                    ec_ws.title = 'ExCov'
                    ec_wb.save(ec_xl_fn)

                    # Run Prism and get outputs for each individual specification
                    ec.run_prism(universes, subsets_arrays, ec_outputs, ec_prism_nt, ec_wb, ec_ws, ec_xl_fn, ec_opt)
                    ec_opt = ec.ec_prism_menu()

            """
            Finished running ExCov
            """
            ec_wb.remove(ec_wb['EC_Template'])
            ec_wb.remove(ec_wb['EC_Prism_Template'])
            ec_wb.save(ec_xl_fn)
            logging.info('Output Excel file is: ' + ec_xl_fn)
            logging.info('Closing workbook')
            ec_wb.close()

            # create MATLAB file
            logging.info('Create MATLAB file')
            misc.ec_create_m_file(Un=universes, Su=subsets_arrays)

        """
        SAT SELECTED
        """
        if problem_type == 3:
            # Prep for excel output
            run_count = 0
            # Setup workbook for data recording
            xl_wb = loadwb(template_dir + 'SAT_Template.xlsx')
            sat_wb_num_of_sheets = len(xl_wb.sheetnames)
            xl_fn = misc.file_name_cformat('SAT_{0}.xlsx')
            # Grab the active worksheet
            xl_ws = xl_wb.active

            # Pick a model checker (NuSMV, nuXmv or PRISM)
            str_modc = misc.modcheck_select()

            # Start output
            print('3-CNF SAT Network Verification')
            logging.info('3-CNF SAT Network Verification')
            main_opt = ''
            while main_opt != 2:  # While not Quit
                logging.info('Printing SAT menu')
                sat.print_sat_menu()
                logging.info('Printed SAT menu.')
                # Get user input for menu selection
                valid_opt = -1
                while valid_opt == -1:
                    main_opt = misc.int_input()
                    if main_opt in [1, 2]:
                        valid_opt = 1
                        print('Selected option ', str(main_opt))
                        logging.info('Selected option: ' + str(main_opt))
                    else:
                        print('Invalid option selected. '
                              'Please select option 1 or 2.')
                        logging.info('Invalid option selected.')

                # If selected sample size:
                if main_opt == 1:
                    xl_wb.save(xl_fn)
                    prompt_samp_size = ('What sample size would you like to'
                                        ' generate?\nINPUT SHOULD BE INTEGER,'
                                        ' i.e. [50]\n')
                    sample_size = misc.int_input(out_str=prompt_samp_size)
                    logging.info('Sample size is: ' + str(sample_size))
                    for row in range(6, sample_size + 6):
                        __ = xl_ws.cell(column=1, row=row, value=(row - 6))

                    # Get max allowed number of variable to randomly generate
                    maxv_p = ('Please enter max number of variables allowed.\n')
                    max_vars = misc.int_input(out_str=maxv_p)

                    if str_modc == "NuSMV" or str_modc == "nuXmv":
                        # Add another worksheet based on the template
                        source = xl_wb['Template']
                        xl_ws = xl_wb.copy_worksheet(source)
                        xl_ws.title = ('Run_' + str(run_count) + '_MaxVars_'
                                       + str(max_vars))
                        run_count += 1

                        # Generate DIMACS samples
                        dimacs_fns, nm_tuples = sat.cnf_gen(sample_size, max_vars,
                                                            xl_ws, xl_wb, xl_fn)

                        logging.info('DIMACS file names are: ' + str(dimacs_fns))
                        logging.info('(Var, Clause) Tuples are: ' + str(nm_tuples))
                        xl_wb.save(xl_fn)

                        # Run DIMACS samples in MiniSat Solver to check satisfiability
                        logging.info('Check satisfiability using MiniSat')
                        minisat_results = sat.mini_sat_solver(dimacs_fns,
                                                              sample_size, xl_ws,
                                                              xl_wb, xl_fn)
                        logging.info('Satisfiability checked')
                        xl_wb.save(xl_fn)
                        # Read each DIMACS, generate two network descriptions
                        # Get list of file names for each
                        logging.info('Generate network descriptions')
                        smv_nc_fns, smv_c_fns = sat.dimacs_to_smv(dimacs_fns,
                                                                  sample_size, xl_ws,
                                                                  xl_wb, xl_fn)
                        logging.info('Network descriptions generated')
                        xl_wb.save(xl_fn)

                        # Run NuSMV on all samples (LTL, CTL, variable re-ordering)
                        logging.info('Run NuSMV on both network descriptions')
                        sat.smv_run_specs(smv_nc_fns, smv_c_fns, sample_size,
                                          xl_ws, xl_wb, xl_fn, str_modc)
                        logging.info('NuSMV runs complete')
                        xl_wb.save(xl_fn)

                    elif str_modc == "prism":
                        # Add another worksheet based on the template
                        source = xl_wb['Template_Prism']
                        xl_ws = xl_wb.copy_worksheet(source)
                        xl_ws.title = ('Run_' + str(run_count) + '_MaxVars_'
                                       + str(max_vars))
                        run_count += 1

                        # Generate DIMACS samples
                        dimacs_fns, nm_tuples = sat.cnf_gen(sample_size, max_vars,
                                                            xl_ws, xl_wb, xl_fn)

                        logging.info('DIMACS file names are: ' + str(dimacs_fns))
                        logging.info('(Var, Clause) Tuples are: ' + str(nm_tuples))
                        xl_wb.save(xl_fn)

                        # Run DIMACS samples in MiniSat Solver to check satisfiability
                        logging.info('Check satisfiability using MiniSat')
                        minisat_results = sat.mini_sat_solver(dimacs_fns,
                                                              sample_size, xl_ws,
                                                              xl_wb, xl_fn)
                        logging.info('Satisfiability checked')
                        xl_wb.save(xl_fn)
                        # Read each DIMACS, generate two network descriptions
                        # Get list of file names for each
                        logging.info('Generate network descriptions')
                        prism_fns = sat.dimacs_to_prism(dimacs_fns,
                                                        sample_size, xl_ws,
                                                        xl_wb, xl_fn)
                        logging.info('Network descriptions generated')
                        xl_wb.save(xl_fn)

                        # Run NuSMV on all samples (LTL, CTL, variable re-ordering)
                        logging.info('Run prism on network descriptions')
                        sat.prism_run_specs(prism_fns, sample_size,
                                            xl_ws, xl_wb, xl_fn, str_modc)
                        logging.info('prism runs complete')
                        xl_wb.save(xl_fn)

                elif main_opt == 2:
                    if len(xl_wb.sheetnames) > sat_wb_num_of_sheets:
                        xl_wb.remove(xl_wb['Template'])
                        xl_wb.remove(xl_wb['Template_Prism'])
                        xl_wb.save(xl_fn)
                        logging.info('Output Excel file is: ' + xl_fn)
                    print('Closing Python Script')

    # Finished running, close logging
    logging.info('Selected Quit')
    logging.info('Closing python script')
    logging.shutdown()
    print('Output directory: ' + cwd)
    print('Log file location: ' + log_path)


def cmd_menu(args):

    # parsing the arguments from command line
    problem_type = misc.cmd_parsing_problem(args.problem)
    ssp_opt = misc.cmd_parsing_opt(args.opt)
    with_tags = misc.cmd_parsing_tags(args.tags)
    str_modc_list = misc.cmd_parsing_mc(args.modecheck)
    mu = misc.cmd_parsing_error(args.error)
    prism_spec = misc.cmd_parsing_prism_spec(args.spec)
    filename = args.filename
    vro = misc.cmd_parsing_vro(args.vro)
    verbosity = misc.cmd_parsing_verbosity(args.verbosity)
    bit_mapping = misc.cmd_parsing_bit_mapping(args.bit_mapping)
    cut = misc.cmd_parsing_cut(args.cut_in_u)
    ic3 = misc.cmd_parsing_ic3(args.ic3)

    """
    MAIN
    -----
    This is the main body of the script for cmd run. All functions are called from here.

    LOGGING
    --------
    Setup log file and logging format
    """
    log_dir = 'Logs'
    if not os.path.exists(log_dir):
        os.mkdir(log_dir)
    log_path = misc.file_name_cformat(os.path.join(os.getcwd(), log_dir, 'log_{0}.log'))
    fmt = '%(asctime)s\t%(module)s\t%(funcName)s\t-\t\t%(levelname)s:\t%(message)s'
    logging.basicConfig(filename=log_path, level=logging.DEBUG, format=fmt,
                        datefmt='%Y-%m-%d %H:%M:%S')

    """
    CURRENT WORKING DIRECTORY
    --------------------------
    Make a directory for output files
    Change the working directory to this new directory
    """
    logging.info('Change working directory')
    dir_name = misc.file_name_cformat('bionetverification_out_{0}')
    os.mkdir(dir_name)
    cwd = os.getcwd()
    template_dir = cwd + '/Template/'
    input_dir = cwd + '/Inputs/'
    new_cwd = os.path.join(cwd, dir_name)
    os.chdir(new_cwd)
    cwd = os.getcwd()
    print('Current working directory:\n' + cwd)
    logging.info('New working directory:' + cwd)

    print('Bionetverification\n')
    logging.info('Bionetverification')

    """
    General Network SELECTED
    """
    if problem_type == 0:
        """
        Get and Parse General Network from input file
        --------------------------------------
        """
        # While filename is not in Inputs
        gn_fn = input_dir + filename
        depth, split_junc, force_down_junc = misc.read_gn(fn=gn_fn)

        # Setup worksheet for data recording
        gn_wb = loadwb(template_dir + 'GN_Template.xlsx')
        gn_wb_num_of_sheets = len(gn_wb.sheetnames)
        gn_xl_fn = misc.file_name_cformat('gn_{0}.xlsx')
        gn_wb.save(gn_xl_fn)

        if str_modc_list[0] == 'NuSMV':
            # Add another worksheet based on the template
            s_source = gn_wb['SINGLE_Template']
            gn_s_ws = gn_wb.copy_worksheet(s_source)
            gn_s_ws.title = ('GN_smv')
            gn_wb.save(gn_xl_fn)

            # generate smv file
            gn_smv_fn = f'GN_depth_{depth}.smv'
            gn.smv_gen(gn_smv_fn, depth, split_junc, force_down_junc)

            # run smv file
            gn.run_nusmv_gn([gn_smv_fn], gn_wb, gn_s_ws, gn_xl_fn, str_modchecker=str_modc_list[0], depth=[depth])

        elif str_modc_list[0] == 'prism':
            # Add another worksheet based on the template
            s_source = gn_wb['Prism_Template']
            gn_s_ws = gn_wb.copy_worksheet(s_source)
            gn_s_ws.title = ('GN_Prism')
            gn_wb.save(gn_xl_fn)

            gn_smv_fn = f'GN_depth_{depth}_mu_{mu}.pm'

            # generate prism file
            gn.prism_gen(gn_smv_fn, depth, split_junc, force_down_junc, mu=mu_user_input)
            if mu > .0:
                gn_smv_fn_no_error = f'GN_depth_{depth}_mu_0.pm'
                gn.prism_gen(gn_smv_fn_no_error, depth, split_junc, force_down_junc, mu=mu_user_input)
                gn_smv_fn_arr = [gn_smv_fn_no_error, gn_smv_fn]
            else:
                gn_smv_fn_arr = [gn_smv_fn]

            # run prism file
            gn.run_prism_gn(gn_smv_fn_arr, gn_wb, gn_s_ws, gn_xl_fn, str_modchecker=str_modc_list[0], depth=[depth], spec_number=prism_spec, mu=mu)

        if len(gn_wb.sheetnames) > gn_wb_num_of_sheets:
            gn_wb.remove(gn_wb['SINGLE_Template'])
            gn_wb.remove(gn_wb['Prism_Template'])
            gn_wb.save(gn_xl_fn)
            logging.info('Output Excel file is: ' + gn_xl_fn)
        else:
            logging.info('There was no action in file: ' + gn_xl_fn)
            logging.info('delete file: ' + gn_xl_fn)
            os.remove(gn_xl_fn)
        logging.info('Closing workbook')
        gn_wb.close()

    """
    SSP SELECTED
    """
    if problem_type == 1:
        for str_modc in str_modc_list:
            """
            Get and Parse SSP sets from input file
            --------------------------------------
            """
            ssp_fn = input_dir + filename
            ssp_arr, num_sets = ssp.read_ssp(ssp_fn)

            # Setup worksheet for data recording
            ssp_wb = loadwb(template_dir + 'SSP_Template.xlsx')
            ssp_wb_num_of_sheets = len(ssp_wb.sheetnames)
            ssp_xl_fn = misc.file_name_cformat(f'SSP_{str_modc}.xlsx')
            ssp_wb.save(ssp_xl_fn)

            if str_modc == "NuSMV" or str_modc == "nuXmv" or str_modc == 'all':
                """
                Generate smv files
                """

                ssp_smv = []
                ssp_smv_nt = []
                ssp_smv_new = []
                ssp_smv_nt_new = []

                # Use specification per output
                if ssp_opt == 1 or ssp_opt == 2:
                    ssp_smv, ssp_smv_nt = ssp.smv_gen(ssp_arr, str_modc, with_tags=with_tags)
                # Use new specifications (csum and nsum for whole network)
                if ssp_opt == 3:
                    ssp_smv_new, ssp_smv_nt_new = ssp.smv_gen_newspec(ssp_arr, str_modc, with_tags=with_tags)

                # If selected bulk run
                if ssp_opt == 1:
                    """
                    Run NuSMV
                    Bulk run specs for per output specs
                    ------------------
                    """
                    # Add another worksheet based on the template
                    a_source = ssp_wb['ALL_Template']
                    ssp_a_ws = ssp_wb.copy_worksheet(a_source)
                    ssp_a_ws.title = 'Bulk_OutSpec'
                    ssp_wb.save(ssp_xl_fn)

                    # Run NuSMV and get output filename for specification
                    ssp.run_nusmv_all(ssp_arr, ssp_smv, ssp_smv_nt, ssp_wb, ssp_a_ws, ssp_xl_fn, str_modc, with_tags=with_tags, ic3=ic3, verbosity=verbosity)

                # If selected individual out run
                elif ssp_opt == 2:
                    """
                    Run NuSMV
                    Run each per output spec individually
                    ------------------
                    """
                    # Add another worksheet based on the template
                    s_source = ssp_wb['SINGLE_Template']
                    ssp_s_ws = ssp_wb.copy_worksheet(s_source)
                    ssp_s_ws.title = 'Single_OutSpec'
                    ssp_wb.save(ssp_xl_fn)

                    # Run NuSMV and get outputs for each individual specification
                    ssp.run_nusmv_single(ssp_arr, ssp_smv, ssp_smv_nt, ssp_wb, ssp_s_ws, ssp_xl_fn, str_modc, with_tags=with_tags, ic3=ic3, verbosity=verbosity)

                # If selected general specifications
                elif ssp_opt == 3:
                    """
                    Run NuSMV
                    Run new specs (csum and nsum for whole network)
                    ------------------
                    """
                    # Add another worksheet based on the template
                    s_source = ssp_wb['NewSpec_Template']
                    ssp_s_ws = ssp_wb.copy_worksheet(s_source)
                    ssp_s_ws.title = 'SSP_GenSpec'
                    ssp_wb.save(ssp_xl_fn)

                    # Run NuSMV and get outputs for each individual specification
                    ssp.run_nusmv_newspec(ssp_arr, ssp_smv_new, ssp_smv_nt_new, ssp_wb, ssp_s_ws, ssp_xl_fn, str_modc, with_tags=with_tags, verbosity=verbosity)

            elif str_modc == "prism":

                """
                Generate prism files
                """
                # Use specification per output
                ssp_prism_nt = ssp.prism_gen(ssp_arr, mu)

                """
                Run Prism
                ------------------
                """
                # Add another worksheet based on the template
                s_source = ssp_wb['Prism_Template']
                ssp_s_ws = ssp_wb.copy_worksheet(s_source)
                ssp_s_ws.title = 'SSP_Prism_Results'
                ssp_wb.save(ssp_xl_fn)

                # Run Prism and get outputs for each individual specification
                ssp.run_prism(ssp_arr, ssp_prism_nt, ssp_wb, ssp_s_ws, ssp_xl_fn, str_modc, prism_spec)

            # If selected return to main
            """
            Finished running SSP problems
            Remove template sheets and close file
            In case there was no action - delete the file
            """
            if len(ssp_wb.sheetnames) > ssp_wb_num_of_sheets:
                ssp_wb.remove(ssp_wb['ALL_Template'])
                ssp_wb.remove(ssp_wb['SINGLE_Template'])
                ssp_wb.remove(ssp_wb['NewSpec_Template'])
                ssp_wb.remove(ssp_wb['Prism_Template'])
                ssp_wb.save(ssp_xl_fn)
                logging.info('Output Excel file is: ' + ssp_xl_fn)
            else:
                logging.info('There was no action in file: ' + ssp_xl_fn)
                logging.info('delete file: ' + ssp_xl_fn)
                os.remove(ssp_xl_fn)
            logging.info('Closing workbook')
            ssp_wb.close()

            # create MATLAB file
            logging.info('Create MATLAB file')
            misc.ssp_create_m_file(Su=ssp_arr)

    """
    ExCov SELECTED
    """
    if problem_type == 2:
        for str_modc in str_modc_list:
            """
            Get and Parse ExCov universes and sets from input file
            --------------------------------------
            """
            excov_fn = input_dir + filename
            universes, subsets_arrays, num_probs = ec.read_ec(excov_fn)

            # Setup worksheet for data recording
            ec_wb = loadwb(template_dir + 'EC_Template.xlsx')
            ec_xl_fn = misc.file_name_cformat(f'ExCov_{str_modc}.xlsx')
            ec_wb.save(ec_xl_fn)

            if str_modc == "NuSMV" or str_modc == "nuXmv":
                """
                Generate smv files
                """
                ec_smv, ec_smv_nt, ec_outputs, max_sums = ec.smv_gen(universes, subsets_arrays, bit_mapping=bit_mapping, with_tags=with_tags, cut_in_u=cut)

                """
                Run NuSMV
                Single spec for exact cover output
                ------------------
                """
                # Add another worksheet based on the template
                source = ec_wb['EC_Template']
                ec_ws = ec_wb.copy_worksheet(source)
                ec_ws.title = 'ExCov'
                ec_wb.save(ec_xl_fn)

                # Run NuSMV and get outputs for each individual specification
                ec.run_nusmv(universes, subsets_arrays, ec_outputs, ec_smv, ec_smv_nt, ec_wb, ec_ws, ec_xl_fn, str_modc, max_sums, with_tags=with_tags, ic3=ic3, verbosity=verbosity)

            elif str_modc == "prism":

                """
                Generate prism files
                """

                ec_prism_nt, ec_outputs, max_sums = ec.prism_gen(universes, subsets_arrays, mu, bit_mapping=bit_mapping)

                """
                Run prism
                Single spec for exact cover output
                ------------------
                """
                # Add another worksheet based on the template
                source = ec_wb['EC_Prism_Template']
                ec_ws = ec_wb.copy_worksheet(source)
                ec_ws.title = 'ExCov'
                ec_wb.save(ec_xl_fn)

                # Run Prism and get outputs for each individual specification
                ec.run_prism(universes, subsets_arrays, ec_outputs, ec_prism_nt, ec_wb, ec_ws, ec_xl_fn, prism_spec)

            """
            Finished running ExCov
            """
            ec_wb.remove(ec_wb['EC_Template'])
            ec_wb.remove(ec_wb['EC_Prism_Template'])
            ec_wb.save(ec_xl_fn)
            logging.info('Output Excel file is: ' + ec_xl_fn)
            logging.info('Closing workbook')
            ec_wb.close()

            # create MATLAB file
            logging.info('Create MATLAB file')
            misc.ec_create_m_file(Un=universes, Su=subsets_arrays)

    """
    SAT SELECTED
    """
    if problem_type == 3:
        for str_modc in str_modc_list:
            # Prep for excel output
            run_count = 0
            # Setup workbook for data recording
            xl_wb = loadwb(template_dir + 'SAT_Template.xlsx')
            sat_wb_num_of_sheets = len(xl_wb.sheetnames)
            xl_fn = misc.file_name_cformat(f'SAT_{str_modc}.xlsx')
            # Grab the active worksheet
            xl_ws = xl_wb.active

            # Start output
            print('3-CNF SAT Network Verification')
            logging.info('3-CNF SAT Network Verification')
            xl_wb.save(xl_fn)

            # Add another worksheet based on the template
            source = None
            if str_modc == 'prism':
                source = xl_wb['Template_Prism']
            else:
                source = xl_wb['Template']
            xl_ws = xl_wb.copy_worksheet(source)
            xl_ws.title = ('Run_' + f'{str_modc}')
            run_count += 1

            # Read list file of dimacs file names
            sat_fn = input_dir + filename
            dimacs_fns = sat.cmd_fn_to_dimacs_fns(input_dir, sat_fn)
            sample_size = len(dimacs_fns)

            # Run DIMACS samples in MiniSat Solver to check satisfiability
            logging.info('Check satisfiability using MiniSat')
            minisat_results = sat.mini_sat_solver(dimacs_fns,
                                                  sample_size, xl_ws,
                                                  xl_wb, xl_fn)

            logging.info('Satisfiability checked')
            xl_wb.save(xl_fn)

            if str_modc == "NuSMV" or str_modc == "nuXmv":

                # Read each DIMACS, generate two network descriptions
                # Get list of file names for each
                logging.info('Generate network descriptions')
                smv_nc_fns, smv_c_fns = sat.dimacs_to_smv(dimacs_fns,
                                                          sample_size, xl_ws,
                                                          xl_wb, xl_fn)
                logging.info('Network descriptions generated')
                xl_wb.save(xl_fn)

                # Run NuSMV on all samples (LTL, CTL, variable re-ordering)
                logging.info(f'Run {str_modc} on both network descriptions')
                sat.smv_run_specs(smv_nc_fns, smv_c_fns, sample_size,
                                  xl_ws, xl_wb, xl_fn, str_modc, vro=vro, verbosity=verbosity)
                logging.info(f'{str_modc} runs complete')
                xl_wb.save(xl_fn)

            elif str_modc == "prism":

                # Get list of file names for each
                logging.info('Generate network descriptions')
                prism_fns = sat.dimacs_to_prism(dimacs_fns,
                                                sample_size, xl_ws,
                                                xl_wb, xl_fn)
                logging.info('Network descriptions generated')
                xl_wb.save(xl_fn)

                # Run Prism on all samples
                logging.info('Run prism on network descriptions')
                sat.prism_run_specs(prism_fns, sample_size,
                                    xl_ws, xl_wb, xl_fn, str_modc)
                logging.info('prism runs complete')
                xl_wb.save(xl_fn)

            if len(xl_wb.sheetnames) > sat_wb_num_of_sheets:
                xl_wb.remove(xl_wb['Template'])
                xl_wb.remove(xl_wb['Template_Prism'])
                xl_wb.save(xl_fn)
                logging.info('Output Excel file is: ' + xl_fn)

    # Finished running, close logging
    logging.info('Selected Quit')
    logging.info('Closing python script')
    logging.shutdown()
    print('Output directory: ' + cwd)
    print('Log file location: ' + log_path)
